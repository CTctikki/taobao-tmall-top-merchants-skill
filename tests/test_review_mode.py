import json
import sys
import tempfile
import unittest
import zipfile
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from audit_review_shops import (
    AuditPaused,
    PageInspection,
    ProductEvidence,
    ReviewTask,
    WebcliBrowserAdapter,
    audit_queue,
    classify_review,
    deduplicate_products,
    make_safe_assign_script,
    stable_identity,
    parse_review_sales_lower_bound,
    summarize_evidence,
)
from review_workbook import (
    ReviewWorkbookError,
    build_review_output,
    compare_workbooks,
    discover_review_tables,
    plan_review_workbook,
    resolve_owner,
    select_review_rows,
)


HEADERS = ["店铺名称", "负责人", "是否符合引入画像", "引入优先级", "店铺链接", "手机号"]


def product(product_id, sales, relevant=True, source_type="shop_home"):
    return ProductEvidence(
        product_id=str(product_id),
        product_url=f"https://item.taobao.com/item.htm?id={product_id}",
        title=f"商品{product_id}",
        sales_text=str(sales),
        sales_lower_bound=sales,
        relevant=relevant,
        source_type=source_type,
        final_page_url="https://shop.taobao.com/search.htm",
    )


def decision(priority="中"):
    relevant_spu = {"高": 20, "中": 15, "低": 10, "待核验": 0}[priority]
    high_sales = {"高": 3, "中": 1, "低": 0, "待核验": 0}[priority]
    return classify_review(relevant_spu, high_sales, complete=priority != "待核验")


def create_book(path, headers=HEADERS, rows=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "审核表"
    sheet.append(headers)
    for row in rows or []:
        sheet.append(row)
    workbook.save(path)
    return path


class FakeAuditor:
    def __init__(self, results):
        self.results = results
        self.calls = []

    def audit(self, task):
        self.calls.append(task.shop_name)
        return self.results[task.shop_name]


class RecordingBrowser:
    def __init__(self, pages=None, failure="", final_url="https://shop.taobao.com/"):
        self.pages = pages or {}
        self.failure = failure
        self.visited = []
        self.assigned_urls = []
        self.sources = []
        self.current_task = None
        self.final_url = final_url

    def set_task(self, task):
        self.current_task = task

    def open_controlled_taobao(self):
        return None

    def assign_url(self, url):
        self.assigned_urls.append(url)

    def inspect_page(self, source_type):
        self.sources.append(source_type)
        if self.failure:
            return PageInspection(
                source_type=source_type,
                requested_url=self.assigned_urls[-1] if self.assigned_urls else "",
                final_url=self.final_url,
                products=(),
                failure=self.failure,
            )
        return PageInspection(
            source_type=source_type,
            requested_url=self.assigned_urls[-1] if self.assigned_urls else "",
            final_url=self.final_url,
            products=tuple(self.pages.get(source_type, [])),
        )

    def open_hot_sales(self, final_url):
        self.assigned_urls.append("https://shop.taobao.com/search.htm?search=y&orderType=hotsell_desc")

    def open_public_shop_list(self, task, final_url):
        self.assigned_urls.append(f"https://shopsearch.taobao.com/search?q={task.shop_name}")


class ReviewEvidenceTests(unittest.TestCase):
    def test_stable_identity_ignores_official_shop_tracking_parameters(self):
        first = stable_identity("甲店", "https://brand.tmall.com/?ali_refid=one&mm_sceneid=1")
        second = stable_identity("甲店", "https://brand.tmall.com/?ali_refid=two&mm_sceneid=2")
        self.assertEqual(first, second)

    def test_parses_public_sales_lower_bounds(self):
        cases = {
            "1万+人付款": 10000,
            "10万+人付款": 100000,
            "总销量：100万+": 1000000,
            "已售1000+件": 1000,
            "9999人付款": 9999,
            "10000人付款": 10000,
        }
        for text, expected in cases.items():
            with self.subTest(text=text):
                self.assertEqual(parse_review_sales_lower_bound(text), expected)

    def test_counts_unique_product_ids_only(self):
        summary = summarize_evidence([product("1", 20000), product("1", 30000), product("2", 9999)])
        self.assertEqual(summary.relevant_spu, 2)
        self.assertEqual(summary.high_sales_links, 1)
        self.assertEqual(deduplicate_products([product("1", 20000), product("1", 30000)])[0].sales_lower_bound, 30000)

    def test_classifies_high_medium_low_and_pending(self):
        self.assertEqual(classify_review(20, 3, complete=True).priority, "高")
        self.assertEqual(classify_review(15, 1, complete=True).priority, "中")
        self.assertEqual(classify_review(14, 5, complete=True).priority, "低")
        self.assertEqual(classify_review(30, 5, complete=False).priority, "待核验")
        self.assertIn("主图能力待人工确认", classify_review(20, 3, complete=True).profile_result)
        self.assertEqual(classify_review(30, 5, complete=False).profile_result, "待核验")


class ReviewWorkbookSelectionTests(unittest.TestCase):
    def test_discovers_shuffled_semantic_headers(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["负责人", "店铺链接", "引入优先级", "店铺名称", "是否符合引入画像"])
        table = discover_review_tables(workbook)[0]
        self.assertEqual(table.columns["shop_name"], 4)
        self.assertEqual(table.columns["profile_result"], 5)
        self.assertEqual(table.columns["priority"], 3)
        self.assertEqual(table.columns["owner"], 1)

    def test_discovers_real_sample_header_phrasing(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "类目",
            "平台",
            "店铺名",
            "是否符合引入画像\n【组内类目 品牌非KA 主图制作能力强 付费（毛利）规模】",
            "引入优先级\n高：SPU≥20；中：SPU≥15；低：SPU<15",
            "辨别采销",
            "店铺联系入口",
            "候选公司（待核验）",
            "公开电话（待核验）",
            "数据来源",
        ])
        table = discover_review_tables(workbook)[0]
        self.assertEqual(table.columns["profile_result"], 4)
        self.assertEqual(table.columns["priority"], 5)
        self.assertEqual(table.columns["owner"], 6)
        self.assertEqual(table.columns["shop_url"], 7)
        self.assertEqual(table.columns["phone"], 9)

    def test_selects_owner_and_only_fully_blank_results(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(HEADERS)
        sheet.append(["甲店", "张萍云", None, None, "https://a", None])
        sheet.append(["乙店", "其他人", None, None, "https://b", None])
        sheet.append(["完成店", "张萍云", "否", "低", "https://c", None])
        sheet.append(["半填店", "张萍云", "否", None, "https://d", None])
        table = discover_review_tables(workbook)[0]
        selection = select_review_rows(workbook, [table], "张萍云")
        self.assertEqual([row.shop_name for row in selection.pending], ["甲店"])
        self.assertEqual([row.shop_name for row in selection.completed], ["完成店"])
        self.assertEqual([row.shop_name for row in selection.inconsistent], ["半填店"])

    def test_excludes_blank_and_merged_subordinate_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(HEADERS)
        sheet.append(["甲店", "张萍云", None, None, "https://a", None])
        sheet.append([None, None, None, None, None, None])
        sheet.merge_cells("A2:A3")
        table = discover_review_tables(workbook)[0]
        selection = select_review_rows(workbook, [table], "张萍云")
        self.assertEqual([row.row_number for row in selection.pending], [2])

    def test_owner_auto_resolution_requires_one_owner(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(HEADERS)
        sheet.append(["甲店", "张萍云", None, None, "https://a", None])
        table = discover_review_tables(workbook)[0]
        self.assertEqual(resolve_owner(workbook, [table]), "张萍云")
        sheet.append(["乙店", "其他人", None, None, "https://b", None])
        with self.assertRaises(ReviewWorkbookError):
            resolve_owner(workbook, [table])


class ReviewBrowserTests(unittest.TestCase):
    def test_long_ad_url_is_serialized_without_shell_splitting(self):
        url = "https://click.simba.taobao.com/x?名称=%E6%B5%8B%E8%AF%95&s=1&k=2&e=3"
        script = make_safe_assign_script(url)
        self.assertIn(json.dumps(url, ensure_ascii=False), script)
        self.assertEqual(script.count("location.assign"), 1)

    def test_webcli_adapter_passes_long_url_only_inside_eval_argument(self):
        url = "https://click.simba.taobao.com/x?名称=%E6%B5%8B%E8%AF%95&s=1&k=2&e=3"
        calls = []
        browser = WebcliBrowserAdapter(session="review", runner=lambda *args: calls.append(args) or {})
        browser.assign_url(url)
        self.assertEqual(calls[0][0:2], ("review", "eval"))
        self.assertIn(json.dumps(url, ensure_ascii=False), calls[0][2])
        self.assertEqual(len(calls[0]), 3)

    def test_falls_back_home_then_hot_sales_then_public_list(self):
        browser = RecordingBrowser(pages={"public_shop_list": [product("1", 10000)]})
        checkpoint = Path(tempfile.mkdtemp()) / "checkpoint.json"
        task = ReviewTask(shop_name="甲店", shop_url="https://click.simba.taobao.com/x?a=1&b=2")
        result = audit_queue([task], browser, checkpoint, sleeper=lambda _: None)[0]
        self.assertEqual(browser.sources, ["shop_home", "shop_hot_sales", "public_shop_list"])
        self.assertEqual(result.evidence[0].source_type, "public_shop_list")

    def test_public_list_fallback_keeps_official_shop_url_from_home(self):
        class PerSourceUrlBrowser(RecordingBrowser):
            def inspect_page(self, source_type):
                inspection = super().inspect_page(source_type)
                urls = {
                    "shop_home": "https://brand.tmall.com/",
                    "shop_hot_sales": "https://brand.tmall.com/search.htm?search=y&orderType=hotsell_desc",
                    "public_shop_list": "https://s.taobao.com/search?q=brand&tab=shop",
                }
                return PageInspection(
                    source_type=inspection.source_type,
                    requested_url=inspection.requested_url,
                    final_url=urls[source_type],
                    products=inspection.products,
                    evidence_complete=inspection.evidence_complete,
                )

        browser = PerSourceUrlBrowser(pages={"public_shop_list": [product("1", 10000)]})
        checkpoint = Path(tempfile.mkdtemp()) / "checkpoint.json"
        result = audit_queue(
            [ReviewTask("甲店", "https://click.simba.taobao.com/x")],
            browser,
            checkpoint,
            sleeper=lambda _: None,
        )[0]
        self.assertEqual(result.official_shop_url, "https://brand.tmall.com/")
        self.assertEqual(result.sources[-1]["final_url"], "https://s.taobao.com/search?q=brand&tab=shop")

    def test_stops_after_home_page_when_products_exist(self):
        browser = RecordingBrowser(pages={"shop_home": [product("1", 10000)]})
        checkpoint = Path(tempfile.mkdtemp()) / "checkpoint.json"
        task = ReviewTask(shop_name="甲店", shop_url="https://shop.taobao.com/a")
        audit_queue([task], browser, checkpoint, sleeper=lambda _: None)
        self.assertEqual(browser.sources, ["shop_home"])
        self.assertEqual(browser.current_task.category, "")

    def test_unclear_category_with_visible_products_is_pending_not_low(self):
        class UnclearCategoryBrowser(RecordingBrowser):
            def inspect_page(self, source_type):
                inspection = super().inspect_page(source_type)
                return PageInspection(
                    source_type=inspection.source_type,
                    requested_url=inspection.requested_url,
                    final_url=inspection.final_url,
                    products=inspection.products,
                    evidence_complete=False,
                )

        browser = UnclearCategoryBrowser(pages={"shop_home": [product("1", 10000, relevant=False)]})
        checkpoint = Path(tempfile.mkdtemp()) / "checkpoint.json"
        result = audit_queue([ReviewTask("甲店", "https://a")], browser, checkpoint, sleeper=lambda _: None)[0]
        self.assertEqual(result.priority, "待核验")
        self.assertEqual(result.profile_result, "待核验")

    def test_category_with_zero_matching_products_is_pending_not_low(self):
        browser = RecordingBrowser(pages={"shop_home": [product("1", 10000, relevant=False)]})
        checkpoint = Path(tempfile.mkdtemp()) / "checkpoint.json"
        result = audit_queue(
            [ReviewTask("甲店", "https://a", category="卫浴用品")],
            browser,
            checkpoint,
            sleeper=lambda _: None,
        )[0]
        self.assertEqual(result.priority, "待核验")

    def test_risk_condition_saves_checkpoint_and_stops_queue(self):
        for failure in (
            "login_required",
            "captcha",
            "slider",
            "hidden_tmd_challenge",
            "bridge_disconnected",
            "navigation_timeout",
            "incompatible_page",
        ):
            with self.subTest(failure=failure), tempfile.TemporaryDirectory() as directory:
                browser = RecordingBrowser(failure=failure)
                checkpoint = Path(directory) / "checkpoint.json"
                tasks = [ReviewTask("甲店", "https://a"), ReviewTask("乙店", "https://b")]
                with self.assertRaises(AuditPaused):
                    audit_queue(tasks, browser, checkpoint, sleeper=lambda _: None)
                saved = json.loads(checkpoint.read_text(encoding="utf-8"))
                self.assertEqual(saved["status"], "paused")
                self.assertEqual(saved["current_task"]["shop_name"], "甲店")
                self.assertEqual(browser.assigned_urls[0], "https://a")
                self.assertNotIn("cookie", json.dumps(saved).lower())

    def test_waits_only_between_shops(self):
        browser = RecordingBrowser(pages={"shop_home": [product("1", 10000)]})
        checkpoint = Path(tempfile.mkdtemp()) / "checkpoint.json"
        waits = []
        audit_queue(
            [ReviewTask("甲店", "https://a"), ReviewTask("乙店", "https://b")],
            browser,
            checkpoint,
            sleeper=waits.append,
            rng=lambda start, end: 19,
        )
        self.assertEqual(waits, [19])

    def test_completed_checkpoint_skips_visit_and_wait(self):
        with tempfile.TemporaryDirectory() as directory:
            checkpoint = Path(directory) / "checkpoint.json"
            browser = RecordingBrowser(pages={"shop_home": [product("1", 10000)]})
            first = ReviewTask("甲店", "https://shop.taobao.com/a")
            audit_queue([first], browser, checkpoint, sleeper=lambda _: None)
            fresh_browser = RecordingBrowser(pages={"shop_home": [product("2", 10000)]})
            waits = []
            results = audit_queue(
                [first], fresh_browser, checkpoint, sleeper=waits.append, rng=lambda start, end: 20
            )
            self.assertEqual(fresh_browser.assigned_urls, [])
            self.assertEqual(waits, [])
            self.assertEqual(results[0].shop_name, "甲店")

    def test_same_name_with_different_official_url_is_not_reused(self):
        with tempfile.TemporaryDirectory() as directory:
            checkpoint = Path(directory) / "checkpoint.json"
            first = ReviewTask("同名店", "https://shop.taobao.com/a")
            audit_queue(
                [first],
                RecordingBrowser(pages={"shop_home": [product("1", 10000)]}),
                checkpoint,
                sleeper=lambda _: None,
            )
            browser = RecordingBrowser(pages={"shop_home": [product("2", 10000)]})
            audit_queue(
                [ReviewTask("同名店", "https://shop.taobao.com/b")],
                browser,
                checkpoint,
                sleeper=lambda _: None,
            )
            self.assertEqual(browser.assigned_urls, ["https://shop.taobao.com/b"])

    def test_same_name_resolving_to_conflicting_official_url_pauses(self):
        with tempfile.TemporaryDirectory() as directory:
            checkpoint = Path(directory) / "checkpoint.json"
            task_a = ReviewTask("同名店", "https://entry.example/a")
            audit_queue(
                [task_a],
                RecordingBrowser(pages={"shop_home": [product("1", 10000)]}, final_url="https://shop.taobao.com/a"),
                checkpoint,
                sleeper=lambda _: None,
            )
            with self.assertRaises(AuditPaused):
                audit_queue(
                    [ReviewTask("同名店", "https://entry.example/b")],
                    RecordingBrowser(pages={"shop_home": [product("2", 10000)]}, final_url="https://shop.taobao.com/b"),
                    checkpoint,
                    sleeper=lambda _: None,
                )
            saved = json.loads(checkpoint.read_text(encoding="utf-8"))
            self.assertEqual(saved["failure"], "shop_identity_conflict")

    def test_reused_first_task_does_not_delay_first_real_visit(self):
        with tempfile.TemporaryDirectory() as directory:
            checkpoint = Path(directory) / "checkpoint.json"
            first = ReviewTask("甲店", "https://shop.taobao.com/a")
            audit_queue(
                [first],
                RecordingBrowser(pages={"shop_home": [product("1", 10000)]}),
                checkpoint,
                sleeper=lambda _: None,
            )
            browser = RecordingBrowser(pages={"shop_home": [product("2", 10000)]})
            waits = []
            audit_queue(
                [first, ReviewTask("乙店", "https://shop.taobao.com/b")],
                browser,
                checkpoint,
                sleeper=waits.append,
                rng=lambda start, end: 20,
            )
            self.assertEqual(browser.assigned_urls, ["https://shop.taobao.com/b"])
            self.assertEqual(waits, [])


class ReviewWorkbookOutputTests(unittest.TestCase):
    def test_invariant_check_detects_hidden_and_print_setting_changes(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            output = Path(directory) / "output.xlsx"
            create_book(source, rows=[["甲店", "张萍云", None, None, "https://a", None]])
            workbook = load_workbook(source)
            sheet = workbook.active
            sheet.print_area = "A1:F2"
            sheet.print_title_rows = "1:1"
            sheet.page_setup.orientation = "landscape"
            hidden = workbook.create_sheet("隐藏说明")
            hidden["A1"] = "说明"
            hidden.sheet_state = "hidden"
            workbook.save(source)
            output.write_bytes(source.read_bytes())

            changed = load_workbook(output)
            changed["隐藏说明"].sheet_state = "visible"
            changed.active.print_area = "A1:D2"
            changed.active.page_setup.orientation = "portrait"
            changed.save(output)

            differences = compare_workbooks(source, output, allowed_cells=set())
            self.assertTrue(any("sheet_state" in item for item in differences))
            self.assertTrue(any("print_area" in item for item in differences))
            self.assertTrue(any("page_setup" in item for item in differences))

    def test_dry_run_plan_groups_duplicates_without_writing_output(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            output = Path(directory) / "output.xlsx"
            create_book(
                source,
                rows=[
                    ["甲店", "张萍云", None, None, "https://a", None],
                    ["甲店", "张萍云", None, None, "https://a", None],
                    ["乙店", "张萍云", "否", "低", "https://b", None],
                ],
            )
            summary = plan_review_workbook(source, owner="张萍云")
            self.assertEqual(summary["pending_rows"], 2)
            self.assertEqual(summary["unique_tasks"], 1)
            self.assertEqual(summary["shops"], ["甲店"])
            self.assertFalse(output.exists())

    def test_duplicate_shop_is_audited_once_and_fills_all_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            output = Path(directory) / "output.xlsx"
            create_book(
                source,
                rows=[
                    ["卡贝官方旗舰店", "张萍云", None, None, "https://a", None],
                    ["卡贝官方旗舰店", "张萍云", None, None, "https://a", None],
                ],
            )
            auditor = FakeAuditor({"卡贝官方旗舰店": decision("中")})
            build_review_output(source, output, owner="张萍云", auditor=auditor)
            result = load_workbook(output)
            sheet = result.active
            self.assertEqual(auditor.calls, ["卡贝官方旗舰店"])
            self.assertEqual(sheet["D2"].value, "中")
            self.assertEqual(sheet["D3"].value, "中")
            self.assertEqual(sheet["C2"].value, sheet["C3"].value)

    def test_existing_duplicate_result_is_reused_without_browser(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            output = Path(directory) / "output.xlsx"
            create_book(
                source,
                rows=[
                    ["甲店", "张萍云", "否", "低", "https://a", None],
                    ["甲店", "张萍云", None, None, "https://a", None],
                ],
            )
            auditor = FakeAuditor({})
            build_review_output(source, output, owner="张萍云", auditor=auditor)
            result = load_workbook(output)
            self.assertEqual(auditor.calls, [])
            self.assertEqual(result.active["C3"].value, "否")
            self.assertEqual(result.active["D3"].value, "低")

    def test_existing_same_name_with_different_direct_url_is_not_reused(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            output = Path(directory) / "output.xlsx"
            create_book(
                source,
                rows=[
                    ["同名店", "张萍云", "否", "低", "https://shop.taobao.com/a", None],
                    ["同名店", "张萍云", None, None, "https://shop.taobao.com/b", None],
                ],
            )
            auditor = FakeAuditor({"同名店": decision("高")})
            build_review_output(source, output, owner="张萍云", auditor=auditor)
            result = load_workbook(output)
            self.assertEqual(auditor.calls, ["同名店"])
            self.assertEqual(result.active["D3"].value, "高")

    def test_source_not_overwritten_and_structure_preserved(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            output = Path(directory) / "output.xlsx"
            create_book(source, rows=[["甲店", "张萍云", None, None, "https://a", 13800138000]])
            workbook = load_workbook(source)
            sheet = workbook.active
            sheet["A2"].hyperlink = "https://shop.taobao.com/a"
            sheet["A2"].font = Font(bold=True)
            sheet["A2"].fill = PatternFill("solid", fgColor="FFFF00")
            sheet["G1"] = "公式"
            sheet["G2"] = "=1+1"
            sheet.freeze_panes = "B2"
            sheet.auto_filter.ref = "A1:G2"
            sheet.column_dimensions["A"].width = 28
            sheet.row_dimensions[2].height = 24
            workbook.save(source)
            original_bytes = source.read_bytes()

            build_review_output(source, output, owner="张萍云", auditor=FakeAuditor({"甲店": decision("低")}))

            self.assertEqual(source.read_bytes(), original_bytes)
            self.assertTrue(zipfile.is_zipfile(output))
            result = load_workbook(output, data_only=False)
            out_sheet = result.active
            self.assertEqual(out_sheet["G2"].value, "=1+1")
            self.assertEqual(out_sheet["A2"].hyperlink.target, "https://shop.taobao.com/a")
            self.assertTrue(out_sheet["A2"].font.bold)
            self.assertEqual(out_sheet["A2"].fill.fgColor.rgb, "00FFFF00")
            self.assertEqual(out_sheet.freeze_panes, "B2")
            self.assertEqual(out_sheet.auto_filter.ref, "A1:G2")
            self.assertEqual(out_sheet.column_dimensions["A"].width, 28)
            self.assertEqual(out_sheet.row_dimensions[2].height, 24)
            self.assertEqual(out_sheet["F2"].value, "13800138000")

    def test_refuses_to_overwrite_source(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            create_book(source, rows=[])
            with self.assertRaises(ReviewWorkbookError):
                build_review_output(source, source, owner="张萍云", auditor=FakeAuditor({}))


if __name__ == "__main__":
    unittest.main()
