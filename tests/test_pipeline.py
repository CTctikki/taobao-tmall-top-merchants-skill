import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from build_workbook import build, candidate_outreach, contact_row_height, extract_contacts, resolve_confirmed_subject, selected_company
from audit_storefronts import qualification_link_script, qualification_page_script, parse_qualification_text
from audit_shops import classify_browser_failure, select_candidates, should_reuse_recorded_failure
from common import BrowserTransientError, classify_item, make_search_script, parse_sales_lower_bound, run_o2
from create_job import create_job
from crosscheck_trademarks import match_trademarks
from verify_job import verify


class PipelineTests(unittest.TestCase):
    def test_default_candidate_audit_limit_is_thirty(self):
        with tempfile.TemporaryDirectory() as directory:
            job = create_job("按摩梳", directory)

            self.assertEqual(job["max_candidate_shops"], 30)

    def test_unconfirmed_single_company_candidate_is_not_selected(self):
        enrichment = {
            "候选店": [
                {
                    "company": "候选公司有限公司",
                    "selected": False,
                }
            ]
        }

        self.assertEqual(selected_company(enrichment, "候选店"), {})

    def test_explicitly_selected_company_is_used(self):
        selected = {
            "company": "已确认公司有限公司",
            "selected": True,
            "evidence_type": "credit_code_match",
            "matched_credit_code": "91310000TEST000001",
            "evidence": "https://example.com/platform-qualification",
            "registration": {
                "企业名称": "已确认公司有限公司",
                "统一社会信用代码": "91310000TEST000001",
            },
        }
        enrichment = {
            "确认店": [
                {"company": "其他候选有限公司", "selected": False},
                selected,
            ]
        }

        self.assertEqual(selected_company(enrichment, "确认店"), selected)

    def test_selected_company_rejects_credit_code_label_without_credit_code(self):
        enrichment = {
            "伪闭环店": [
                {
                    "company": "名称相似公司有限公司",
                    "selected": True,
                    "evidence_type": "credit_code_match",
                }
            ]
        }

        self.assertEqual(selected_company(enrichment, "伪闭环店"), {})

    def test_selected_company_rejects_unproven_credit_code_match(self):
        enrichment = {
            "错码店": [
                {
                    "company": "错误公司有限公司",
                    "selected": True,
                    "evidence_type": "credit_code_match",
                    "matched_credit_code": "91310000DIFFERENT01",
                    "evidence": "https://example.com/platform-qualification",
                    "registration": {
                        "企业名称": "错误公司有限公司",
                        "统一社会信用代码": "91310000COMPANY0001",
                    },
                }
            ]
        }

        self.assertEqual(selected_company(enrichment, "错码店"), {})

    def test_selected_company_rejects_platform_label_outside_qualification_record(self):
        enrichment = {
            "伪资质店": [
                {
                    "company": "伪资质公司有限公司",
                    "selected": True,
                    "evidence_type": "platform_qualification",
                    "evidence": "https://example.com/platform-qualification",
                    "registration": {
                        "企业名称": "伪资质公司有限公司",
                        "统一社会信用代码": "91310000FAKE000001",
                    },
                }
            ]
        }

        self.assertEqual(selected_company(enrichment, "伪资质店"), {})

    def test_verified_qualification_requires_company_and_credit_code(self):
        qualification = {
            "status": "verified",
            "company_name": "缺码公司有限公司",
            "credit_code": "",
            "evidence_type": "platform_qualification",
        }

        self.assertEqual(resolve_confirmed_subject({}, "缺码店", qualification), {})

    def test_selected_company_rejects_non_closing_evidence(self):
        enrichment = {
            "品牌店": [
                {
                    "company": "品牌权利公司有限公司",
                    "selected": True,
                    "evidence_type": "trademark_or_official_site",
                    "subject_role": "商标权利主体",
                }
            ]
        }

        self.assertEqual(selected_company(enrichment, "品牌店"), {})

    def test_parses_tmall_platform_qualification(self):
        text = """
        天猫网店经营者相关资质信息
        企业注册号：91330201MA284E2D7K
        企业名称：义乌市昊杜缝纫电子商务有限公司
        类型：有限责任公司(自然人投资或控股)
        住所：浙江省金华市义乌市北苑街道柳三村柳青二区2幢6单元402
        法定代表人：吴宝峰
        成立时间：2017-02-21
        注册资本：100万元人民币
        登记机关：义乌市市场监督管理局
        """

        result = parse_qualification_text(text, "https://zhaoshang.tmall.com/maintaininfo/liangzhao.htm")

        self.assertEqual(result["status"], "verified")
        self.assertEqual(result["company_name"], "义乌市昊杜缝纫电子商务有限公司")
        self.assertEqual(result["credit_code"], "91330201MA284E2D7K")
        self.assertEqual(result["legal_person"], "吴宝峰")
        self.assertEqual(result["established"], "2017-02-21")
        self.assertIn("liangzhao.htm", result["source_url"])

    def test_storefront_audit_scripts_find_and_guard_qualification_page(self):
        link_script = qualification_link_script()
        page_script = qualification_page_script()

        self.assertIn("查看商家公示信息", link_script)
        self.assertIn("liangzhao.htm", link_script)
        self.assertIn("mouseover", link_script)
        self.assertIn("_____tmd__", page_script)
        self.assertIn("企业名称", page_script)
        self.assertIn("企业注册号", page_script)

    def test_platform_license_overrides_mismatched_selected_candidate(self):
        with tempfile.TemporaryDirectory() as directory:
            job_dir = Path(directory)
            create_job("电动车挡风", job_dir)
            audit = {
                "aijia": {
                    "category": "电动车挡风",
                    "shop_name": "艾佳运动户外旗舰店",
                    "platform": "天猫",
                    "exact_shop_spu_seen": 20,
                    "target_spu": 12,
                    "electric_spu": 0,
                    "accessory_spu": 0,
                    "unrelated_spu": 8,
                    "target_share": 0.6,
                    "passes_minimum": True,
                    "match_grade": "高匹配",
                    "shop_url": "https://shop356249658.taobao.com/category.htm",
                    "user_id": "1",
                    "target_items": [{"sales": "100+人付款"}],
                }
            }
            qualifications = {
                "艾佳运动户外旗舰店": {
                    "status": "verified",
                    "company_name": "义乌市昊杜缝纫电子商务有限公司",
                    "credit_code": "91330201MA284E2D7K",
                    "legal_person": "吴宝峰",
                    "address": "浙江省金华市义乌市北苑街道柳三村柳青二区2幢6单元402",
                    "established": "2017-02-21",
                    "source_url": "https://zhaoshang.tmall.com/maintaininfo/liangzhao.htm",
                    "evidence_type": "platform_qualification",
                }
            }
            enrichment = {
                "艾佳运动户外旗舰店": [
                    {
                        "company": "东莞市艾佳E健身管理有限公司",
                        "selected": True,
                        "evidence_type": "company_name_similarity",
                        "registration": {
                            "企业名称": "东莞市艾佳E健身管理有限公司",
                            "统一社会信用代码": "WRONG-CODE",
                            "注册地址": "广东省东莞市",
                        },
                    },
                    {
                        "company": "深圳市艾佳怡科技有限公司",
                        "selected": False,
                        "registration": {"企业名称": "深圳市艾佳怡科技有限公司"},
                    },
                ]
            }
            (job_dir / "assortment_audit.json").write_text(json.dumps(audit, ensure_ascii=False), encoding="utf-8")
            (job_dir / "platform_qualifications.json").write_text(json.dumps(qualifications, ensure_ascii=False), encoding="utf-8")
            (job_dir / "company_enrichment.json").write_text(json.dumps(enrichment, ensure_ascii=False), encoding="utf-8")

            workbook = build(job_dir)
            loaded = load_workbook(workbook, data_only=False)
            sheet = loaded["正式招商商家"]
            headers = {cell.value: cell.column for cell in sheet[3]}

            self.assertEqual(sheet.cell(4, headers["公司名称"]).value, "义乌市昊杜缝纫电子商务有限公司")
            self.assertEqual(sheet.cell(4, headers["统一社会信用代码"]).value, "91330201MA284E2D7K")
            self.assertEqual(sheet.cell(4, headers["主体一致性"]).value, "平台营业执照已确认")
            self.assertIn("东莞市艾佳E健身管理有限公司", sheet.cell(4, headers["建联候选公司（非店铺主体，待核验）"]).value)
            self.assertNotIn("东莞市艾佳E健身管理有限公司", sheet.cell(4, headers["公司名称"]).value)
            self.assertTrue(verify(job_dir, workbook)["ok"])

            sheet.cell(4, headers["公司名称"]).value = "东莞市艾佳E健身管理有限公司"
            loaded.save(workbook)
            with self.assertRaisesRegex(AssertionError, "正式主体与平台营业执照不一致"):
                verify(job_dir, workbook)

    def test_candidate_outreach_exposes_contacts_without_confirming_subject(self):
        enrichment = {
            "候选店": [
                {
                    "company": "候选公司有限公司",
                    "selected": False,
                    "registration": {
                        "企业名称": "候选公司有限公司",
                        "注册地址": "上海市候选路1号",
                    },
                    "contact": {
                        "联系方式信息": {
                            "电话": [{"电话号码": "021-12345678"}],
                            "邮箱": [{"邮箱": "contact@example.com"}],
                        }
                    },
                }
            ]
        }

        outreach = candidate_outreach(enrichment, "候选店")

        self.assertEqual(selected_company(enrichment, "候选店"), {})
        self.assertIn("候选公司有限公司", outreach["candidate_companies"])
        self.assertIn("021-12345678", outreach["candidate_phones"])
        self.assertIn("contact@example.com", outreach["candidate_emails"])
        self.assertIn("上海市候选路1号", outreach["candidate_addresses"])
        self.assertIn("先核验", outreach["outreach_note"])

    def test_extract_contacts_keeps_all_unique_values(self):
        contact = {
            "联系方式信息": {
                "电话": [
                    {"电话号码": f"1380000000{index}"}
                    for index in range(8)
                ]
                + [{"电话号码": "13800000000"}],
                "邮箱": [
                    {"邮箱": f"contact{index}@example.com"}
                    for index in range(7)
                ]
                + [{"邮箱": "contact0@example.com"}],
            }
        }

        phones, emails = extract_contacts(contact)

        self.assertEqual(len(phones.split("；")), 8)
        self.assertEqual(len(emails.split("；")), 7)
        self.assertEqual(phones.split("；")[0], "13800000000")
        self.assertEqual(emails.split("；")[0], "contact0@example.com")

    def test_contact_row_height_expands_for_long_lists(self):
        phones = "；".join(f"1380000000{index}" for index in range(8))
        emails = "；".join(f"contact{index}@example.com" for index in range(7))

        self.assertEqual(contact_row_height("", "", 68), 68)
        self.assertGreaterEqual(contact_row_height(phones, emails, 68), 120)

    def test_massage_comb_profile_keeps_taobao_and_electric_products(self):
        with tempfile.TemporaryDirectory() as directory:
            job = create_job("按摩梳", directory)
            self.assertFalse(job["exclude_electric"])
            self.assertTrue(classify_item("电动红光头皮按摩梳", job)["relevant"])
            self.assertFalse(classify_item("宠物狗狗按摩梳", job)["relevant"])
            self.assertFalse(classify_item("按摩梳替换头配件", job)["relevant"])

    def test_sales_lower_bound(self):
        self.assertEqual(parse_sales_lower_bound("1.2万人付款"), 12000)
        self.assertEqual(parse_sales_lower_bound("800+人付款"), 800)
        self.assertIsNone(parse_sales_lower_bound(""))

    def test_trademark_crosscheck_requires_brand_and_relevant_class(self):
        marks = [
            {"商标名称": "RAFFINI", "国际分类": "21类 厨房洁具"},
            {"商标名称": "RAFFINI", "国际分类": "33类 酒"},
            {"商标名称": "OTHER", "国际分类": "21类 厨房洁具"},
        ]
        self.assertEqual(
            match_trademarks(marks, ["raffini"], ["8", "20", "21"]),
            [marks[0]],
        )

    def test_tiered_audit_recovers_low_exposure_c_stores(self):
        with tempfile.TemporaryDirectory() as directory:
            job = create_job("按摩梳", directory)
            rows = [
                {"shop_name": "大综合百货", "discovered_target_spu": 1, "queries": ["按摩梳"], "items": [{"sales": "2万+人付款"}]},
                {"shop_name": "叶梳匠品牌店", "discovered_target_spu": 1, "queries": ["按摩梳"], "items": [{"sales": "2万+人付款"}]},
                {"shop_name": "低露出品牌店", "discovered_target_spu": 1, "queries": ["a", "b", "c", "d"], "items": [{"sales": "1.5万+人付款"}]},
            ]
            selected = select_candidates(job, rows)
            self.assertEqual({row["shop_name"] for row in selected}, {"叶梳匠品牌店", "低露出品牌店"})

    def test_quality_shortlist_filters_and_ranks_by_payment_lower_bound(self):
        with tempfile.TemporaryDirectory() as directory:
            job = create_job("按摩梳", directory)
            rows = [
                {"shop_name": "高销量专业店", "discovered_target_spu": 5, "queries": ["a", "b", "c"], "items": [{"sales": "5万+人付款"}]},
                {"shop_name": "中销量专业店", "discovered_target_spu": 4, "queries": ["a", "b", "c"], "items": [{"sales": "3万+人付款"}]},
                {"shop_name": "低销量铺货店", "discovered_target_spu": 20, "queries": ["a", "b", "c", "d"], "items": [{"sales": "100+人付款"}]},
                {"shop_name": "天猫超市", "discovered_target_spu": 30, "queries": ["a", "b", "c", "d"], "items": [{"sales": "20万+人付款"}]},
            ]

            selected = select_candidates(job, rows)

            self.assertEqual([row["shop_name"] for row in selected], ["高销量专业店", "中销量专业店"])
            self.assertEqual(selected[0]["payment_lower_bound"], 50000)

    def test_sales_top_n_mode_keeps_single_query_high_sales_shop(self):
        with tempfile.TemporaryDirectory() as directory:
            job = create_job("按摩梳", directory)
            job.update(
                {
                    "sales_top_n_mode": True,
                    "minimum_payment_lower_bound": 0,
                    "minimum_quality_query_coverage": 1,
                    "max_candidate_shops": 30,
                }
            )
            rows = [
                {"shop_name": "单词高销量店", "discovered_target_spu": 1, "queries": ["按摩梳"], "items": [{"sales": "10万+人付款"}]},
                {"shop_name": "多词普通店", "discovered_target_spu": 3, "queries": ["a", "b", "c"], "items": [{"sales": "1万+人付款"}]},
            ]

            selected = select_candidates(job, rows)

            self.assertEqual([row["shop_name"] for row in selected], ["单词高销量店", "多词普通店"])
            self.assertEqual(selected[0]["audit_reason"], "sales_top")

    @patch("common.time.sleep", return_value=None)
    @patch("common.subprocess.run")
    def test_browser_transient_abort_retries_without_retrying_risk_control(self, run, _sleep):
        aborted = type("Result", (), {"returncode": 1, "stdout": "", "stderr": "This operation was aborted"})()
        success = type("Result", (), {"returncode": 0, "stdout": '{"ok": true}', "stderr": ""})()
        run.side_effect = [aborted, success]
        self.assertEqual(run_o2("session", "eval", "1+1"), {"ok": True})
        self.assertEqual(run.call_count, 2)

    @patch("common.time.sleep", return_value=None)
    @patch("common.subprocess.run")
    def test_browser_timeout_raises_structured_transient_error(self, run, _sleep):
        import subprocess

        run.side_effect = subprocess.TimeoutExpired(["o2", "browser"], 1)
        with self.assertRaises(BrowserTransientError) as context:
            run_o2("session", "eval", "script containing 验证码 text", timeout=1)
        self.assertEqual(str(context.exception), "webcli timed out after 1 seconds")

    def test_search_script_detects_hidden_challenge_and_bounds_mtop_wait(self):
        script = make_search_script("按摩梳", 2)
        self.assertIn("_____tmd__", script)
        self.assertIn("TAOBAO_RISK_CONTROL", script)
        self.assertIn("Promise.race", script)
        self.assertIn("MTOP_REQUEST_TIMEOUT", script)

    def test_risk_control_is_recorded_separately_from_transient_failures(self):
        self.assertEqual(classify_browser_failure("Error: TAOBAO_RISK_CONTROL"), "risk_control")
        self.assertEqual(classify_browser_failure("webcli timed out after 90 seconds"), "browser_transient_error")

    def test_recorded_risk_control_is_reused_without_overwriting_error(self):
        errors = {
            "风控店": {"status": "risk_control", "reason": "TAOBAO_RISK_CONTROL"},
            "瞬时错误店": {"status": "browser_transient_error", "reason": "timeout"},
        }

        self.assertTrue(should_reuse_recorded_failure(errors, "风控店"))
        self.assertFalse(should_reuse_recorded_failure(errors, "瞬时错误店"))
        self.assertFalse(should_reuse_recorded_failure(errors, "新店"))

    def test_workbook_contract(self):
        with tempfile.TemporaryDirectory() as directory:
            job_dir = Path(directory)
            create_job("按摩梳", job_dir)
            audit = {
                "淘宝优质店": {
                    "category": "按摩梳",
                    "shop_name": "淘宝优质店",
                    "platform": "淘宝",
                    "exact_shop_spu_seen": 20,
                    "target_spu": 10,
                    "electric_spu": 2,
                    "accessory_spu": 1,
                    "unrelated_spu": 9,
                    "target_share": 0.5,
                    "passes_minimum": True,
                    "match_grade": "高匹配",
                    "shop_url": "https://shop.example/taobao",
                    "user_id": "1",
                    "target_items": [{"sales": "100+人付款"}],
                },
                "天猫达标店": {
                    "category": "采耳工具",
                    "shop_name": "天猫达标店",
                    "platform": "天猫",
                    "exact_shop_spu_seen": 30,
                    "target_spu": 10,
                    "electric_spu": 0,
                    "accessory_spu": 2,
                    "unrelated_spu": 18,
                    "target_share": 1 / 3,
                    "passes_minimum": True,
                    "match_grade": "达标",
                    "shop_url": "https://shop.example/tmall",
                    "user_id": "2",
                    "target_items": [{"sales": "80+人付款"}],
                },
            }
            (job_dir / "assortment_audit.json").write_text(json.dumps(audit, ensure_ascii=False), encoding="utf-8")
            (job_dir / "storefronts.json").write_text("{}", encoding="utf-8")
            (job_dir / "company_candidates.json").write_text("{}", encoding="utf-8")
            (job_dir / "company_enrichment.json").write_text(
                json.dumps(
                    {
                        "淘宝优质店": [
                            {
                                "company": "候选公司有限公司",
                                "selected": False,
                                "registration": {
                                    "企业名称": "候选公司有限公司",
                                    "注册地址": "上海市候选路1号",
                                },
                                "contact": {
                                    "联系方式信息": {
                                        "电话": [{"电话号码": "021-12345678"}],
                                        "邮箱": [{"邮箱": "contact@example.com"}],
                                    }
                                },
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            (job_dir / "audit_errors.json").write_text(
                json.dumps(
                    {
                        "风控待续跑店": {
                            "status": "risk_control",
                            "reason": "TAOBAO_RISK_CONTROL",
                            "target": {"platform": "淘宝", "shop_url": "https://shop.example/risk"},
                        }
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            workbook = build(job_dir)
            loaded = load_workbook(workbook, data_only=False)
            formal_categories = {
                loaded["正式招商商家"].cell(row, 2).value
                for row in range(4, loaded["正式招商商家"].max_row + 1)
            }
            self.assertEqual(formal_categories, {"按摩梳", "采耳工具"})
            headers = {cell.value: cell.column for cell in loaded["正式招商商家"][3]}
            self.assertFalse(loaded["正式招商商家"].cell(4, headers["公司名称"]).value)
            self.assertIn("021-12345678", loaded["正式招商商家"].cell(4, headers["候选电话（待核验）"]).value)
            self.assertIn("上海市候选路1号", loaded["正式招商商家"].cell(4, headers["候选地址（待核验）"]).value)
            unresolved_values = [cell.value for row in loaded["未确认字段"].iter_rows() for cell in row]
            self.assertIn("风控待续跑店", unresolved_values)
            self.assertIn("店铺商品结构审计", unresolved_values)
            result = verify(job_dir, workbook)
            self.assertEqual(result["formal_records"], 2)
            self.assertEqual(result["platforms"], ["天猫", "淘宝"])


    def test_workbook_keeps_trademark_owner_as_candidate_and_labels_exact_search_failure(self):
        with tempfile.TemporaryDirectory() as directory:
            job_dir = Path(directory)
            create_job("\u6309\u6469\u68b3", job_dir)
            audit = {
                "selected-shop": {
                    "category": "\u6309\u6469\u68b3",
                    "shop_name": "\u5929\u732b\u8fbe\u6807\u5e97",
                    "platform": "\u5929\u732b",
                    "exact_shop_spu_seen": 20,
                    "target_spu": 10,
                    "electric_spu": 0,
                    "accessory_spu": 0,
                    "unrelated_spu": 10,
                    "target_share": 0.5,
                    "passes_minimum": True,
                    "match_grade": "\u9ad8\u5339\u914d",
                    "shop_url": "https://shop.example/tmall",
                    "user_id": "2",
                    "target_items": [{"sales": "80+\u4eba\u4ed8\u6b3e"}],
                }
            }
            (job_dir / "assortment_audit.json").write_text(json.dumps(audit, ensure_ascii=False), encoding="utf-8")
            (job_dir / "company_enrichment.json").write_text(
                json.dumps(
                    {
                        "\u5929\u732b\u8fbe\u6807\u5e97": [
                            {
                                "company": "\u54c1\u724c\u6743\u5229\u516c\u53f8\u6709\u9650\u516c\u53f8",
                                "selected": True,
                                "subject_role": "\u5546\u6807\u6743\u5229\u4e3b\u4f53",
                                "subject_confidence": "\u9ad8\uff08\u54c1\u724c\u5546\u6807+\u76f8\u5173\u7c7b\u522b\uff09",
                                "pending": "\u5f53\u524d\u5929\u732b\u6301\u8bc1\u8fd0\u8425\u4e3b\u4f53\u4ecd\u5f85\u5e73\u53f0\u8d44\u8d28\u9875\u6700\u7ec8\u786e\u8ba4",
                                "registration": {"\u4f01\u4e1a\u540d\u79f0": "\u54c1\u724c\u6743\u5229\u516c\u53f8\u6709\u9650\u516c\u53f8"},
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            (job_dir / "audit_errors.json").write_text(
                json.dumps(
                    {
                        "\u7cbe\u786e\u53cd\u67e5\u672a\u547d\u4e2d\u5e97": {
                            "status": "search_no_exact_shop_items",
                            "reason": "100 search results contained no exact shop-name match",
                            "target": {"platform": "\u6dd8\u5b9d", "shop_url": "https://shop.example/no-match"},
                        }
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            workbook = build(job_dir)
            loaded = load_workbook(workbook, data_only=False)
            headers = {cell.value: cell.column for cell in loaded["\u6b63\u5f0f\u62db\u5546\u5546\u5bb6"][3]}
            self.assertIsNone(loaded["\u6b63\u5f0f\u62db\u5546\u5546\u5bb6"].cell(4, headers["\u516c\u53f8\u540d\u79f0"]).value)
            self.assertIn("\u54c1\u724c\u6743\u5229\u516c\u53f8\u6709\u9650\u516c\u53f8", loaded["\u6b63\u5f0f\u62db\u5546\u5546\u5bb6"].cell(4, headers["\u5efa\u8054\u5019\u9009\u516c\u53f8\uff08\u975e\u5e97\u94fa\u4e3b\u4f53\uff0c\u5f85\u6838\u9a8c\uff09"]).value)
            self.assertEqual(loaded["\u6b63\u5f0f\u62db\u5546\u5546\u5bb6"].cell(4, headers["\u4e3b\u4f53\u4e00\u81f4\u6027"]).value, "\u672a\u786e\u8ba4")
            self.assertIn("\u4f01\u4e1a\u641c\u7d22\u7ed3\u679c\u53ea\u80fd\u4f5c\u4e3a\u5f85\u6838\u9a8c\u5efa\u8054\u5019\u9009", loaded["\u6b63\u5f0f\u62db\u5546\u5546\u5bb6"].cell(4, headers["\u5f85\u786e\u8ba4\u9879"]).value)
            unresolved_values = [cell.value for row in loaded["\u672a\u786e\u8ba4\u5b57\u6bb5"].iter_rows() for cell in row]
            self.assertTrue(any("\u641c\u7d22\u53cd\u67e5\u672a\u547d\u4e2d\u7cbe\u786e\u5e97\u94fa\u5546\u54c1" in str(value) for value in unresolved_values))


class SkillMetadataTests(unittest.TestCase):
    def test_skill_has_no_secret_and_correct_name(self):
        text = (ROOT / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("name: taobao-tmall-top-merchants", text)
        self.assertIn("淘宝/天猫top商家清单", text)
        self.assertIn("_____tmd__", text)
        self.assertIn("MTOP_REQUEST_TIMEOUT", text)
        self.assertIn("查看商家公示信息", text)
        self.assertIn("liangzhao.htm", text)
        self.assertIn("FN_API_KEY", text)
        self.assertIn("biz_fuzzy_search", text)
        self.assertIn("biz_basic_info", text)
        self.assertIn("联系方式不能单独证明店铺主体", text)
        self.assertIn("selected: false", text)
        self.assertIn("风鸟模糊发现 → 企查查精确核验 → 风鸟补缺", text)
        self.assertIn("企查查精确核验 → 风鸟补缺", text)
        self.assertIn("平台资质页 > 信用代码一致 > 商标/品牌官网 > 企业名称相似 > 电话邮箱", text)
        self.assertNotIn("Bearer MX", text)
        self.assertIn("scripts/bootstrap.ps1", text)
        self.assertIn("--install-missing", text)

    def test_skill_supports_user_provided_shop_list_mode(self):
        text = (ROOT / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("用户指定名单模式", text)
        for input_type in ["店铺名", "多个链接", "文本", "表格", "电子表格", "混合信息"]:
            self.assertIn(input_type, text)
        self.assertIn("不运行 `mine_taobao.py`、`audit_shops.py`", text)
        self.assertIn("全部进入正式招商商家", text)
        self.assertIn("不得补造目标SPU、店内目标商品占比、付款人数展示下限或Top30结论", text)
        self.assertIn("用户指定名单，不代表Top30或主营准入达标", text)
        self.assertIn("名单模式跳过第3至第7步", text)
        self.assertIn("不运行 `create_job.py`、`mine_taobao.py`、`audit_shops.py`、`audit_storefronts.py`", text)

    def test_skill_routes_independent_review_mode(self):
        skill = (ROOT / "SKILL.md").read_text(encoding="utf-8")
        readme = (ROOT / "README.md").read_text(encoding="utf-8")
        metadata = (ROOT / "agents" / "openai.yaml").read_text(encoding="utf-8")
        review = (ROOT / "references" / "review-mode.md").read_text(encoding="utf-8")
        contract = (ROOT / "references" / "review-data-contract.md").read_text(encoding="utf-8")

        self.assertIn("审核筛选模式", skill)
        self.assertIn("不得并入用户指定名单模式", skill)
        self.assertIn("scripts/review_workbook.py", skill)
        self.assertIn("references/review-mode.md", skill)
        self.assertIn("-AuditOnly", readme)
        self.assertIn("审核", metadata)
        for phrase in ["不要求企查查", "不要求风鸟", "唯一商品 ID", "18–22 秒", "待核验"]:
            self.assertIn(phrase, review)
        for phrase in ["stable_identity", "profile_result", "high_sales_links", "source_type"]:
            self.assertIn(phrase, contract)
        self.assertNotIn("近30天销量", review)

    def test_skill_requires_both_private_keys_and_guides_browser_bridge_setup(self):
        skill = (ROOT / "SKILL.md").read_text(encoding="utf-8")
        readme = (ROOT / "README.md").read_text(encoding="utf-8")
        setup = (ROOT / "references" / "mcp-setup.md").read_text(encoding="utf-8")
        combined = "\n".join([skill, readme, setup])

        self.assertIn("https://agent.qcc.com/profile/api-key", combined)
        self.assertIn("https://www.riskbird.com/center/apiKey", combined)
        self.assertIn("必须同时提供", skill)
        self.assertIn("缺一不可", skill)
        self.assertIn("公共额度不能替代", skill)
        self.assertIn("configure_enterprise_keys.py", skill)
        self.assertIn("run_fengniao.py", skill)
        self.assertIn("标准输入", skill)
        self.assertIn("不要让用户自行配置环境变量", skill)
        for instruction in ["chrome://extensions", "开发者模式", "加载已解压的扩展程序", ".webcli/extension", "固定", "保持 Chrome 开启"]:
            self.assertIn(instruction, combined)

    def test_skill_requires_platform_license_to_match_confirmed_subject(self):
        skill = (ROOT / "SKILL.md").read_text(encoding="utf-8")
        contract = (ROOT / "references" / "data-contract.md").read_text(encoding="utf-8")

        self.assertIn("platform_qualifications.json", skill)
        self.assertIn("平台营业执照主体必须与正式表", skill)
        self.assertIn("建联候选公司（非店铺主体，待核验）", skill)
        self.assertIn("商标/品牌官网证据不得单独写入正式主体", skill)
        self.assertIn("matched_credit_code", skill)
        self.assertIn("平台资质只能来自独立的 platform_qualifications.json", skill)
        self.assertIn("platform_qualifications.json", contract)
        self.assertIn("company_name", contract)
        self.assertIn("credit_code", contract)
        self.assertIn("matched_credit_code", contract)


if __name__ == "__main__":
    unittest.main()
