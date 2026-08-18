import argparse
import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common import load_job, parse_sales_lower_bound


NAVY = "17365D"
BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
GREEN = "E2F0D9"
YELLOW = "FFF2CC"
RED = "F4CCCC"
WHITE = "FFFFFF"
TEXT = "1F2937"
THIN = Side(style="thin", color="D6DCE4")


def load_optional(path, default):
    return json.loads(path.read_text(encoding="utf-8")) if path.exists() else default


def extract_contacts(contact):
    info = contact.get("联系方式信息", {}) if isinstance(contact, dict) else {}
    phones = [item.get("电话号码", "") for item in info.get("电话", []) if item.get("电话号码")]
    emails = [item.get("邮箱", "") for item in info.get("邮箱", []) if item.get("邮箱")]
    return "；".join(dict.fromkeys(phones)), "；".join(dict.fromkeys(emails))


def contact_row_height(phones, emails, minimum):
    phone_lines = len(phones.split("；")) if phones else 0
    email_lines = len(emails.split("；")) if emails else 0
    return max(minimum, 16 * max(phone_lines, email_lines) + 8)


def sales_summary(audit):
    values = [parse_sales_lower_bound(item.get("sales")) for item in audit.get("target_items", [])]
    values = [value for value in values if value is not None]
    return sum(values), len(values)


def selected_company(enrichment, shop):
    candidates = enrichment.get(shop, [])
    selected = next((row for row in candidates if row.get("selected") in {True, "是", "yes"}), None)
    return selected or (candidates[0] if len(candidates) == 1 else None) or {}


def prepare_audit_failures(job_dir, job):
    errors = load_optional(job_dir / "audit_errors.json", {})
    labels = {
        "risk_control": "淘宝风控挑战（需人工验证）",
        "browser_transient_error": "Browser Bridge/MTop瞬时超时",
        "browser_error": "浏览器错误",
        "skipped_by_operator": "持续异常后显式跳过",
    }
    actions = {
        "risk_control": "用户完成淘宝滑块/验证码后从断点续跑",
        "browser_transient_error": "恢复Bridge后换新会话低频补跑；出现风控立即停止",
        "browser_error": "先运行webcli doctor诊断，再按错误类型补跑",
        "skipped_by_operator": "恢复后移除--skip-shop并定向补跑",
    }
    failures = []
    for shop_name, record in errors.items():
        target = record.get("target", {})
        status = record.get("status", "browser_error")
        first_line = str(record.get("reason", "")).splitlines()[0]
        reason = labels.get(status, status)
        if first_line and first_line not in reason:
            reason = f"{reason}；{first_line}"
        failures.append({
            "category": job["category"],
            "platform": target.get("platform", ""),
            "shop_name": shop_name,
            "company": "",
            "missing_fields": "店铺商品结构审计",
            "reason": reason,
            "next_step": actions.get(status, "核对错误账本后定向补跑"),
            "store_url": target.get("shop_url", ""),
        })
    return failures


def prepare_rows(job_dir):
    job = load_job(job_dir)
    audit = load_optional(job_dir / "assortment_audit.json", {})
    storefronts = load_optional(job_dir / "storefronts.json", {})
    enrichment = load_optional(job_dir / "company_enrichment.json", {})
    formal, rejected = [], []
    for record in audit.values():
        store = storefronts.get(record["shop_name"], {})
        if not record.get("passes_minimum"):
            reasons = []
            if record.get("target_spu", 0) < job["min_spu"]:
                reasons.append(f'目标SPU<{job["min_spu"]}')
            if record.get("target_share", 0) < job["min_share"]:
                reasons.append(f'目标占比<{job["min_share"]:.0%}')
            rejected.append({**record, "reason": "；".join(reasons), "store_url": store.get("official_url") or record.get("shop_url", "")})
            continue
        company_record = selected_company(enrichment, record["shop_name"])
        registration = company_record.get("registration", {}) if isinstance(company_record, dict) else {}
        if not isinstance(registration, dict) or registration.get("error"):
            registration = {}
        phone, email = extract_contacts(company_record.get("contact", {})) if company_record else ("", "")
        payment, sales_count = sales_summary(record)
        company = registration.get("企业名称") or company_record.get("company", "")
        formal.append({
            **record,
            "shop_type": store.get("shop_type") or ("天猫店" if record["platform"] == "天猫" else "淘宝店（企业/个人待核验）"),
            "store_url": store.get("official_url") or record.get("shop_url", ""),
            "shop_id": store.get("shop_id", ""),
            "seller_id": store.get("seller_id", ""),
            "payment_lower_bound": payment,
            "sales_item_count": sales_count,
            "company": company,
            "legal_person": registration.get("法定代表人", ""),
            "phone": phone,
            "email": email,
            "address": registration.get("注册地址", ""),
            "established": registration.get("成立日期", ""),
            "credit_code": registration.get("统一社会信用代码", ""),
            "status": registration.get("登记状态", ""),
            "subject_role": company_record.get("role", "") if company_record else "",
            "subject_confidence": company_record.get("confidence", "未确认") if company_record else "未确认",
            "evidence": company_record.get("evidence", "") if company_record else "",
            "pending": "人工打开店铺资质页确认当前持证主体" if company else "主体待店铺资质页确认；多候选未自动选择",
        })
    formal.sort(key=lambda row: (0 if row["match_grade"] == "高匹配" else 1, -row["payment_lower_bound"], row["shop_name"]))
    rejected.sort(key=lambda row: (-row.get("target_spu", 0), row["shop_name"]))
    missing = []
    fields = [("公司名称", "company"), ("法人", "legal_person"), ("公司电话", "phone"), ("邮箱", "email"), ("注册地址", "address"), ("成立日期", "established"), ("统一社会信用代码", "credit_code")]
    for row in formal:
        absent = [label for label, key in fields if not row.get(key)]
        if absent:
            missing.append({"category": job["category"], "platform": row["platform"], "shop_name": row["shop_name"], "company": row["company"], "missing_fields": "、".join(absent), "reason": "主体未唯一确认" if not row["company"] else "公开渠道未披露或本次MCP未返回", "next_step": row["pending"], "store_url": row["store_url"]})
    return job, formal, rejected, missing, prepare_audit_failures(job_dir, job)


def style_title(sheet, title, subtitle, end_column):
    end_letter = get_column_letter(end_column)
    sheet.merge_cells(f"A1:{end_letter}1")
    sheet["A1"] = title
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color=WHITE)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells(f"A2:{end_letter}2")
    sheet["A2"] = subtitle
    sheet["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sheet["A2"].font = Font(name="Microsoft YaHei", size=10, color=TEXT)
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 42


def style_table(sheet, header_row, widths, row_height=48):
    border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    for cell in sheet[header_row]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=9, color=TEXT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        sheet.row_dimensions[row[0].row].height = row_height
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    sheet.sheet_view.showGridLines = False


def add_overview(workbook, job, formal, rejected, missing, audit_failures):
    sheet = workbook.active
    sheet.title = "概览"
    style_title(sheet, f'{job["category"]}｜淘宝 + 天猫TOP商家招商概览', f'正式门槛：目标SPU≥{job["min_spu"]}、店内目标占比≥{job["min_share"]:.0%}；≥{job["high_match_share"]:.0%}标记高匹配。淘宝/C店不是淘汰条件。', 8)
    sheet.append(["指标", "总计", "天猫", "淘宝", "高匹配", "达标", "主体已补", "存在未确认字段"])
    sheet.append(["正式招商记录", len(formal), sum(row["platform"] == "天猫" for row in formal), sum(row["platform"] == "淘宝" for row in formal), sum(row["match_grade"] == "高匹配" for row in formal), sum(row["match_grade"] == "达标" for row in formal), sum(bool(row["company"]) for row in formal), len(missing)])
    sheet.append(["淘汰记录", len(rejected), "", "", "", "", "", ""])
    sheet.append(["未完成审计", len(audit_failures), sum(row["platform"] == "天猫" for row in audit_failures), sum(row["platform"] == "淘宝" for row in audit_failures), "", "", "", "详见未确认字段"])
    sheet.append([])
    sheet.append(["口径", "说明", "", "", "", "", "", ""])
    notes = [("类目范围", job.get("scope_note", "")), ("精准命中", "按精确店铺名反聚合商品，并以唯一商品ID计SPU。"), ("销量", "付款人数展示下限用于相对排序，不等同近30天月销。"), ("主体", "公司可能是品牌/生产/运营候选，最终以店铺资质页为准；多候选不自动取第一名。")]
    for label, text in notes:
        sheet.append([label, text, "", "", "", "", "", ""])
        sheet.merge_cells(start_row=sheet.max_row, start_column=2, end_row=sheet.max_row, end_column=8)
    style_table(sheet, 3, [22, 16, 14, 14, 14, 14, 16, 20], 38)


def add_formal(workbook, job, rows):
    sheet = workbook.create_sheet("正式招商商家")
    headers = ["序号", "类目", "平台/店铺类型", "店铺名", "目标SPU", "精确店铺SPU", "相关占比", "匹配等级", "付款人数展示下限", "有销量展示商品数", "店铺链接", "shopId", "sellerId", "主体角色", "主体置信度", "公司名称", "法人", "公司电话", "邮箱", "注册地址", "成立日期", "统一社会信用代码", "登记状态", "数据来源/证据", "待确认项"]
    style_title(sheet, f'{job["category"]}｜正式招商商家｜淘宝 + 天猫', f'保留规则：目标SPU≥{job["min_spu"]}、相关占比≥{job["min_share"]:.0%}。公司字段尽量补齐，但持证运营主体仍需平台资质页终审。', len(headers))
    sheet.append(headers)
    for index, row in enumerate(rows, 1):
        excel_row = sheet.max_row + 1
        sheet.append([index, job["category"], f'{row["platform"]}｜{row["shop_type"]}', row["shop_name"], row["target_spu"], row["exact_shop_spu_seen"], f"=IFERROR(E{excel_row}/F{excel_row},0)", f'=IF(AND(E{excel_row}>={job["min_spu"]},G{excel_row}>={job["high_match_share"]}),"高匹配",IF(AND(E{excel_row}>={job["min_spu"]},G{excel_row}>={job["min_share"]}),"达标","不达标"))', row["payment_lower_bound"], row["sales_item_count"], row["store_url"], row["shop_id"], row["seller_id"], row["subject_role"], row["subject_confidence"], row["company"], row["legal_person"], row["phone"], row["email"], row["address"], row["established"], row["credit_code"], row["status"], row["evidence"], row["pending"]])
        sheet.cell(excel_row, 7).number_format = "0.0%"
    style_table(sheet, 3, [7, 12, 24, 25, 11, 13, 11, 11, 16, 16, 40, 14, 16, 20, 28, 30, 12, 26, 28, 44, 13, 24, 20, 42, 38], 68)
    for row_number, row in enumerate(rows, 4):
        sheet.row_dimensions[row_number].height = contact_row_height(row["phone"], row["email"], 68)
    if rows:
        sheet.conditional_formatting.add(f"G4:G{sheet.max_row}", CellIsRule(operator="greaterThanOrEqual", formula=[str(job["high_match_share"])], fill=PatternFill("solid", fgColor=GREEN)))
        sheet.conditional_formatting.add(f"G4:G{sheet.max_row}", CellIsRule(operator="between", formula=[str(job["min_share"]), str(job["high_match_share"] - 0.000001)], fill=PatternFill("solid", fgColor=YELLOW)))
        for row_number in range(4, sheet.max_row + 1):
            if not sheet.cell(row_number, 16).value:
                sheet.cell(row_number, 16).fill = PatternFill("solid", fgColor=RED)


def add_subjects(workbook, job_dir, formal):
    sheet = workbook.create_sheet("主体核验")
    headers = ["店铺名", "平台", "候选类型", "公司/主体", "状态", "成立日期", "统一社会信用代码", "法人", "电话", "邮箱", "关联证据", "核验结论"]
    style_title(sheet, "主体核验｜企业候选与工商补全", "企查查/爱企查/天眼查/风鸟等多源检索返回多候选时不自动选择第一名；正式采用项仍需店铺资质页确认。", len(headers))
    sheet.append(headers)
    platform = {row["shop_name"]: row["platform"] for row in formal}
    candidates = load_optional(job_dir / "company_candidates.json", {})
    enrichment = load_optional(job_dir / "company_enrichment.json", {})
    for shop, record in candidates.items():
        for company in record.get("result", {}).get("企业信息", [])[:8]:
            legal = company.get("法定代表人名称", [])
            sheet.append([shop, platform.get(shop, ""), "实体识别候选", company.get("企业名称", ""), company.get("状态", ""), company.get("成立日期", ""), company.get("统一社会信用代码", ""), "；".join(legal) if isinstance(legal, list) else legal, "", "", f'检索词：{record.get("query", "")}', "多候选，仅供人工核对"])
    for shop, records in enrichment.items():
        for record in records:
            registration = record.get("registration", {}) if isinstance(record.get("registration"), dict) else {}
            phone, email = extract_contacts(record.get("contact", {}))
            sheet.append([shop, platform.get(shop, ""), "已查工商候选", registration.get("企业名称") or record.get("company", ""), registration.get("登记状态", ""), registration.get("成立日期", ""), registration.get("统一社会信用代码", ""), registration.get("法定代表人", ""), phone, email, record.get("evidence", ""), "正式表暂采用；仍待资质页确认" if record.get("selected") in {True, "是", "yes"} else "候选主体"])
    style_table(sheet, 3, [26, 10, 20, 32, 18, 13, 24, 12, 26, 28, 42, 34], 58)
    for row_number in range(4, sheet.max_row + 1):
        sheet.row_dimensions[row_number].height = contact_row_height(
            sheet.cell(row_number, 9).value,
            sheet.cell(row_number, 10).value,
            58,
        )


def add_missing(workbook, formal, missing, audit_failures):
    sheet = workbook.create_sheet("未确认字段")
    style_title(sheet, "未确认字段汇总｜查不到的数据与下一步方案", "空白不是没有数据，而是本轮没有足够证据安全归属到店铺。优先取得店铺资质页主体全称或信用代码。", 8)
    sheet.append(["字段", "正式记录数", "已取得数", "缺失数", "建议动作", "", "", ""])
    fields = [("公司名称", "company"), ("法人", "legal_person"), ("公司电话", "phone"), ("邮箱", "email"), ("注册地址", "address"), ("成立日期", "established"), ("统一社会信用代码", "credit_code")]
    for label, key in fields:
        obtained = sum(bool(row.get(key)) for row in formal)
        sheet.append([label, len(formal), obtained, len(formal) - obtained, "平台资质页确认主体后再用企业MCP精确补查", "", "", ""])
        sheet.merge_cells(start_row=sheet.max_row, start_column=5, end_row=sheet.max_row, end_column=8)
    detail_header = sheet.max_row + 2
    for column, value in enumerate(["类目", "平台", "店铺名", "当前候选公司", "未确认字段", "原因", "建议动作", "店铺链接"], 1):
        sheet.cell(detail_header, column, value)
    for row in missing + audit_failures:
        sheet.append([row[key] for key in ["category", "platform", "shop_name", "company", "missing_fields", "reason", "next_step", "store_url"]])
    style_table(sheet, 3, [12, 10, 26, 30, 34, 42, 44, 42], 52)
    for cell in sheet[detail_header]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    sheet.auto_filter.ref = f"A{detail_header}:H{sheet.max_row}"
    sheet.freeze_panes = f"A{detail_header + 1}"


def add_rejected(workbook, job, rows):
    sheet = workbook.create_sheet("淘汰商家")
    headers = ["类目", "平台", "店铺名", "目标SPU", "精确店铺SPU", "相关占比", "带电SPU", "配件SPU", "无关SPU", "淘汰原因", "店铺链接", "采集时间"]
    style_title(sheet, "淘汰商家｜低于门槛", "淘汰仅依据本轮商品结构；淘宝/C店不会因店铺类型被淘汰。", len(headers))
    sheet.append(headers)
    for row in rows:
        sheet.append([job["category"], row["platform"], row["shop_name"], row["target_spu"], row["exact_shop_spu_seen"], row["target_share"], row["electric_spu"], row["accessory_spu"], row["unrelated_spu"], row["reason"], row["store_url"], row.get("captured_at", "")])
        sheet.cell(sheet.max_row, 6).number_format = "0.0%"
    style_table(sheet, 3, [12, 10, 26, 11, 13, 12, 11, 11, 11, 28, 42, 24], 48)


def add_method(workbook, job):
    sheet = workbook.create_sheet("口径与复用")
    style_title(sheet, "口径与复用方法｜换其他二级类目也可执行", "先按商品类目发现店铺，再反查店铺商品结构，最后锚定工商主体。", 6)
    sheet.append(["步骤", "目的", "输入", "处理规则", "输出", "质量控制"])
    rows = [(1, "类目发现", "类目核心词+同义词", "淘宝全平台搜索，淘宝与天猫都保留", "候选店铺池", "低频访问"), (2, "店铺反聚合", "候选店铺名", "精确店铺名+唯一商品ID", "精确店铺SPU", "排除其他店铺"), (3, "目标识别", "商品标题", "目标词与排除词", "目标SPU", "抽查误判"), (4, "结构门槛", "目标SPU/店铺SPU", f'SPU≥{job["min_spu"]}且占比≥{job["min_share"]:.0%}', "正式/淘汰", "两平台同规则"), (5, "主体锚定", "店铺/品牌/资质", "实体识别后再查工商", "公司和联系方式", "多候选不自动选"), (6, "人工终审", "店铺资质页", "核对当前持证主体", "最终招商主体", "验证码由用户完成")]
    for row in rows:
        sheet.append(row)
    sheet.append([])
    sheet.append(["参数", "当前值", "说明", "", "", ""])
    params = [("类目", job["category"], job.get("scope_note", "")), ("查询词", "、".join(job["queries"]), "发现词可扩展"), ("目标词", "、".join(job["include_patterns"]), "标题至少命中一个"), ("排除词", "、".join(job["exclude_patterns"] + job["accessory_patterns"]), "配件和无关商品"), ("带电商品", "排除" if job.get("exclude_electric") else "保留", "按类目任务配置")]
    for label, value, note in params:
        sheet.append([label, value, note, "", "", ""])
        sheet.merge_cells(start_row=sheet.max_row, start_column=3, end_row=sheet.max_row, end_column=6)
    style_table(sheet, 3, [10, 22, 30, 48, 28, 34], 54)


def build(job_dir, output=None):
    job_dir = Path(job_dir).resolve()
    job, formal, rejected, missing, audit_failures = prepare_rows(job_dir)
    workbook = Workbook()
    workbook.properties.creator = "taobao-tmall-top-merchants"
    add_overview(workbook, job, formal, rejected, missing, audit_failures)
    add_formal(workbook, job, formal)
    add_subjects(workbook, job_dir, formal)
    add_missing(workbook, formal, missing, audit_failures)
    add_rejected(workbook, job, rejected)
    add_method(workbook, job)
    for sheet in workbook.worksheets:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
        sheet.page_setup.fitToWidth = 2 if sheet.max_column >= 15 else 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_title_rows = "1:3"
    output_path = Path(output) if output else job_dir / "outputs" / f'{job["category"]}_淘宝天猫TOP商家招商表.xlsx'
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--job-dir", required=True)
    parser.add_argument("--output")
    args = parser.parse_args()
    print(build(args.job_dir, args.output).resolve())


if __name__ == "__main__":
    main()
