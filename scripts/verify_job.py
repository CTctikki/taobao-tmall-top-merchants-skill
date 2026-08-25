import argparse
import json
import re
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from common import load_job


SECRET_PATTERN = re.compile(r"Bearer\s+[A-Za-z0-9._-]{12,}|Authorization\s*[=:]", re.I)


def normalize_identity(value):
    return re.sub(r"[\s（）()·•,，。]", "", str(value or "")).lower()


def verify(job_dir, workbook_path=None):
    job_dir = Path(job_dir).resolve()
    job = load_job(job_dir)
    audit = json.loads((job_dir / "assortment_audit.json").read_text(encoding="utf-8"))
    formal_records = [row for row in audit.values() if row.get("passes_minimum")]
    for row in formal_records:
        assert row["target_spu"] >= job["min_spu"], row["shop_name"]
        assert row["target_share"] >= job["min_share"], row["shop_name"]
    workbook_path = Path(workbook_path) if workbook_path else job_dir / "outputs" / f'{job["category"]}_淘宝天猫TOP商家招商表.xlsx'
    with ZipFile(workbook_path) as archive:
        bad = archive.testzip()
        assert bad is None, bad
    formula_book = load_workbook(workbook_path, data_only=False)
    value_book = load_workbook(workbook_path, data_only=True)
    required = ["概览", "正式招商商家", "主体核验", "未确认字段", "淘汰商家", "口径与复用"]
    assert formula_book.sheetnames == required, formula_book.sheetnames
    errors = []
    for sheet in value_book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}:
                    errors.append((sheet.title, cell.coordinate, cell.value))
    assert not errors, errors
    qualifications_path = job_dir / "platform_qualifications.json"
    if qualifications_path.exists():
        qualifications = json.loads(qualifications_path.read_text(encoding="utf-8"))
        formal_sheet = value_book["正式招商商家"]
        headers = {cell.value: cell.column for cell in formal_sheet[3]}
        required_subject_columns = {"店铺名", "公司名称", "统一社会信用代码", "主体一致性"}
        assert required_subject_columns.issubset(headers), "正式招商商家缺少平台主体核验列"
        rows = {formal_sheet.cell(row, headers["店铺名"]).value: row for row in range(4, formal_sheet.max_row + 1)}
        for shop, qualification in qualifications.items():
            if qualification.get("status") != "verified" or shop not in rows:
                continue
            assert qualification.get("company_name") and qualification.get("credit_code"), f"{shop}: verified平台资质缺少公司名或信用代码"
            row = rows[shop]
            actual_company = formal_sheet.cell(row, headers["公司名称"]).value
            actual_code = formal_sheet.cell(row, headers["统一社会信用代码"]).value
            assert normalize_identity(actual_company) == normalize_identity(qualification.get("company_name")), f"{shop}: 正式主体与平台营业执照不一致"
            assert str(actual_code or "").strip().upper() == str(qualification.get("credit_code") or "").strip().upper(), f"{shop}: 信用代码与平台营业执照不一致"
            assert formal_sheet.cell(row, headers["主体一致性"]).value == "平台营业执照已确认", f"{shop}: 主体一致性状态错误"
    for path in job_dir.rglob("*"):
        if path.is_file() and path.suffix.lower() in {".json", ".txt", ".log", ".md", ".py"}:
            text = path.read_text(encoding="utf-8", errors="ignore")
            assert not SECRET_PATTERN.search(text), f"possible secret in {path}"
    return {"ok": True, "category": job["category"], "formal_records": len(formal_records), "platforms": sorted({row["platform"] for row in formal_records}), "workbook": str(workbook_path), "formula_errors": 0}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--job-dir", required=True)
    parser.add_argument("--workbook")
    args = parser.parse_args()
    print(json.dumps(verify(args.job_dir, args.workbook), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
