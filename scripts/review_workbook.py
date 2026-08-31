import argparse
import json
import re
import shutil
import sys
import zipfile
from dataclasses import dataclass, field, replace
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from audit_review_shops import (
    AuditPaused,
    ReviewDecision,
    ReviewTask,
    ShopAuditResult,
    WebcliBrowserAdapter,
    audit_queue,
    normalize_shop_name,
    normalize_url,
)


REQUIRED_FIELDS = ("shop_name", "profile_result", "priority", "owner")
HEADER_ALIASES = {
    "shop_name": {"店铺", "店铺名", "店铺名称", "商家", "商家名称"},
    "profile_result": {"是否符合引入画像", "符合引入画像", "引入画像", "审核结论", "是否符合画像"},
    "priority": {"引入优先级", "优先级", "招商优先级"},
    "owner": {"负责人", "招商负责人", "所属负责人", "辨别采销"},
    "shop_url": {"店铺链接", "店铺网址", "店铺url", "链接", "店铺联系入口"},
    "category": {"类目", "目标类目", "主营类目", "审核类目"},
    "phone": {"手机号", "手机号码", "联系电话", "电话", "公开电话"},
}


class ReviewWorkbookError(RuntimeError):
    pass


@dataclass(frozen=True)
class ReviewTable:
    sheet_name: str
    header_row: int
    columns: dict[str, int]


@dataclass(frozen=True)
class ReviewRow:
    sheet_name: str
    row_number: int
    shop_name: str
    owner: str
    shop_url: str
    category: str
    profile_result: str = ""
    priority: str = ""


@dataclass
class ReviewSelection:
    pending: list[ReviewRow] = field(default_factory=list)
    completed: list[ReviewRow] = field(default_factory=list)
    inconsistent: list[ReviewRow] = field(default_factory=list)


@dataclass(frozen=True)
class ReviewRunSummary:
    output_path: str
    owner: str
    pending_rows: int
    unique_tasks: int
    reused_rows: int
    audited_shops: int
    inconsistent_rows: tuple[tuple[str, int], ...]


def _text(value):
    return str(value).strip() if value is not None else ""


def _normalize_header(value):
    text = _text(value).lower().replace("\n", "").replace("\r", "")
    return re.sub(r"[\s:：()（）\-_]+", "", text)


NORMALIZED_ALIASES = {
    field_name: {_normalize_header(alias) for alias in aliases}
    for field_name, aliases in HEADER_ALIASES.items()
}
PREFIX_FIELDS = {"profile_result", "priority", "phone"}


def _map_header_row(sheet, row_number):
    columns = {}
    for cell in sheet[row_number]:
        normalized = _normalize_header(cell.value)
        if not normalized:
            continue
        for field_name, aliases in NORMALIZED_ALIASES.items():
            matched = normalized in aliases or (
                field_name in PREFIX_FIELDS and any(normalized.startswith(alias) for alias in aliases)
            )
            if matched:
                if field_name in columns:
                    raise ReviewWorkbookError(
                        f"{sheet.title} row {row_number} has duplicate semantic column: {field_name}"
                    )
                columns[field_name] = cell.column
    return columns


def discover_review_tables(workbook, scan_rows=30):
    tables = []
    incomplete = []
    for sheet in workbook.worksheets:
        for row_number in range(1, min(sheet.max_row, scan_rows) + 1):
            columns = _map_header_row(sheet, row_number)
            if all(field_name in columns for field_name in REQUIRED_FIELDS):
                tables.append(ReviewTable(sheet.title, row_number, columns))
                break
            if "shop_name" in columns and len(columns) >= 2:
                incomplete.append((sheet.title, row_number, columns))
    if not tables:
        detail = f"; possible headers={incomplete}" if incomplete else ""
        raise ReviewWorkbookError(f"no review table with required semantic columns{detail}")
    return tables


def _row_value(sheet, row_number, columns, field_name):
    column = columns.get(field_name)
    if not column:
        return ""
    cell = sheet.cell(row_number, column)
    if isinstance(cell, MergedCell):
        return ""
    return _text(cell.value)


def _iter_rows(workbook, tables):
    for table in tables:
        sheet = workbook[table.sheet_name]
        for row_number in range(table.header_row + 1, sheet.max_row + 1):
            shop_name = _row_value(sheet, row_number, table.columns, "shop_name")
            if not shop_name:
                continue
            yield ReviewRow(
                sheet_name=table.sheet_name,
                row_number=row_number,
                shop_name=shop_name,
                owner=_row_value(sheet, row_number, table.columns, "owner"),
                shop_url=_row_value(sheet, row_number, table.columns, "shop_url"),
                category=_row_value(sheet, row_number, table.columns, "category"),
                profile_result=_row_value(sheet, row_number, table.columns, "profile_result"),
                priority=_row_value(sheet, row_number, table.columns, "priority"),
            )


def resolve_owner(workbook, tables, requested_owner=None):
    if _text(requested_owner):
        return _text(requested_owner)
    owners = {
        row.owner for row in _iter_rows(workbook, tables)
        if row.owner and not row.profile_result and not row.priority
    }
    if len(owners) == 1:
        return owners.pop()
    if not owners:
        raise ReviewWorkbookError("no pending rows with a non-empty owner")
    raise ReviewWorkbookError(f"multiple owners found; pass --owner: {sorted(owners)}")


def select_review_rows(workbook, tables, owner):
    selection = ReviewSelection()
    normalized_owner = _normalize_header(owner)
    for row in _iter_rows(workbook, tables):
        if _normalize_header(row.owner) != normalized_owner:
            continue
        if row.profile_result and row.priority:
            selection.completed.append(row)
        elif row.profile_result or row.priority:
            selection.inconsistent.append(row)
        else:
            selection.pending.append(row)
    return selection


def group_pending_rows(rows):
    grouped = {}
    for row in rows:
        key = normalize_shop_name(row.shop_name)
        grouped.setdefault(key, []).append(row)
    tasks = []
    for grouped_rows in grouped.values():
        first = grouped_rows[0]
        urls = {row.shop_url for row in grouped_rows if row.shop_url}
        if len(urls) > 1:
            non_ad_urls = {url for url in urls if "click.simba.taobao.com" not in url.lower()}
            if len(non_ad_urls) > 1:
                raise ReviewWorkbookError(f"conflicting shop URLs for {first.shop_name}")
        tasks.append(
            ReviewTask(
                shop_name=first.shop_name,
                shop_url=first.shop_url,
                category=first.category,
                source_rows=tuple((row.sheet_name, row.row_number) for row in grouped_rows),
            )
        )
    return tasks


def plan_review_workbook(source, owner=None, category=""):
    source_path = Path(source).resolve()
    workbook = load_workbook(
        source_path,
        data_only=False,
        keep_vba=source_path.suffix.lower() == ".xlsm",
    )
    tables = discover_review_tables(workbook)
    selected_owner = resolve_owner(workbook, tables, owner)
    selection = select_review_rows(workbook, tables, selected_owner)
    tasks = group_pending_rows(selection.pending)
    if category:
        tasks = [replace(task, category=task.category or category) for task in tasks]
    return {
        "source": str(source_path),
        "owner": selected_owner,
        "tables": [
            {"sheet": table.sheet_name, "header_row": table.header_row, "columns": table.columns}
            for table in tables
        ],
        "pending_rows": len(selection.pending),
        "unique_tasks": len(tasks),
        "shops": [task.shop_name for task in tasks],
        "inconsistent_rows": [
            {"sheet": row.sheet_name, "row": row.row_number, "shop_name": row.shop_name}
            for row in selection.inconsistent
        ],
    }


def _table_by_sheet(tables):
    return {table.sheet_name: table for table in tables}


def _completed_results(rows):
    results = {}
    for row in rows:
        key = (normalize_shop_name(row.shop_name), normalize_url(row.shop_url))
        value = (row.profile_result, row.priority)
        if key in results and results[key] != value:
            raise ReviewWorkbookError(f"conflicting completed results for {row.shop_name}")
        results[key] = value
    return results


def _decision_values(result):
    if isinstance(result, (ReviewDecision, ShopAuditResult)):
        return result.profile_result, result.priority
    if isinstance(result, dict):
        return _text(result.get("profile_result")), _text(result.get("priority"))
    raise ReviewWorkbookError(f"unsupported auditor result: {type(result).__name__}")


def _snapshot_cell(cell):
    hyperlink = cell.hyperlink.target if cell.hyperlink else ""
    return {
        "value": cell.value,
        "data_type": cell.data_type,
        "style": cell.style_id,
        "number_format": cell.number_format,
        "hyperlink": hyperlink,
    }


def workbook_snapshot(path):
    path = Path(path)
    workbook = load_workbook(path, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    snapshot = {"sheetnames": tuple(workbook.sheetnames), "sheets": {}}
    for sheet in workbook.worksheets:
        cells = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None or cell.has_style or cell.hyperlink:
                    cells[cell.coordinate] = _snapshot_cell(cell)
        snapshot["sheets"][sheet.title] = {
            "cells": cells,
            "sheet_state": sheet.sheet_state,
            "merged": tuple(str(item) for item in sheet.merged_cells.ranges),
            "freeze": str(sheet.freeze_panes or ""),
            "filter": str(sheet.auto_filter.ref or ""),
            "print_area": str(sheet.print_area or ""),
            "print_title_rows": str(sheet.print_title_rows or ""),
            "print_title_cols": str(sheet.print_title_cols or ""),
            "page_setup": (
                sheet.page_setup.orientation,
                sheet.page_setup.paperSize,
                sheet.page_setup.scale,
                sheet.page_setup.fitToWidth,
                sheet.page_setup.fitToHeight,
                sheet.page_setup.firstPageNumber,
                sheet.page_setup.useFirstPageNumber,
                sheet.page_setup.blackAndWhite,
                sheet.page_setup.draft,
                sheet.page_setup.cellComments,
                sheet.page_setup.errors,
                sheet.page_setup.horizontalDpi,
                sheet.page_setup.verticalDpi,
            ),
            "page_margins": (
                sheet.page_margins.left,
                sheet.page_margins.right,
                sheet.page_margins.top,
                sheet.page_margins.bottom,
                sheet.page_margins.header,
                sheet.page_margins.footer,
            ),
            "print_options": (
                sheet.print_options.horizontalCentered,
                sheet.print_options.verticalCentered,
                sheet.print_options.headings,
                sheet.print_options.gridLines,
                sheet.print_options.gridLinesSet,
            ),
            "row_dimensions": {
                index: (dimension.height, dimension.hidden, dimension.outlineLevel)
                for index, dimension in sheet.row_dimensions.items()
            },
            "column_dimensions": {
                index: (dimension.width, dimension.hidden, dimension.outlineLevel)
                for index, dimension in sheet.column_dimensions.items()
            },
        }
    return snapshot


def compare_workbooks(source, output, allowed_cells):
    before = workbook_snapshot(source)
    after = workbook_snapshot(output)
    errors = []
    if before["sheetnames"] != after["sheetnames"]:
        errors.append("worksheet order changed")
    for sheet_name in before["sheetnames"]:
        left = before["sheets"][sheet_name]
        right = after["sheets"].get(sheet_name)
        if right is None:
            errors.append(f"missing worksheet: {sheet_name}")
            continue
        for field_name in (
            "sheet_state",
            "merged",
            "freeze",
            "filter",
            "print_area",
            "print_title_rows",
            "print_title_cols",
            "page_setup",
            "page_margins",
            "print_options",
            "row_dimensions",
            "column_dimensions",
        ):
            if left[field_name] != right[field_name]:
                errors.append(f"{sheet_name} {field_name} changed")
        coordinates = set(left["cells"]) | set(right["cells"])
        for coordinate in coordinates:
            if (sheet_name, coordinate) in allowed_cells:
                continue
            if left["cells"].get(coordinate) != right["cells"].get(coordinate):
                errors.append(f"{sheet_name}!{coordinate} changed")
    return errors


def _phone_cells_to_text(workbook, tables):
    changed = set()
    for table in tables:
        phone_column = table.columns.get("phone")
        if not phone_column:
            continue
        sheet = workbook[table.sheet_name]
        for row_number in range(table.header_row + 1, sheet.max_row + 1):
            cell = sheet.cell(row_number, phone_column)
            if isinstance(cell.value, int) and len(str(cell.value)) >= 7:
                cell.value = str(cell.value)
                cell.number_format = "@"
                changed.add((sheet.title, cell.coordinate))
    return changed


def _write_result(workbook, table, row_number, profile_result, priority):
    sheet = workbook[table.sheet_name]
    profile_cell = sheet.cell(row_number, table.columns["profile_result"])
    priority_cell = sheet.cell(row_number, table.columns["priority"])
    if profile_cell.value not in (None, "") or priority_cell.value not in (None, ""):
        raise ReviewWorkbookError(f"refusing to overwrite existing result at {table.sheet_name}!{row_number}")
    profile_cell.value = profile_result
    priority_cell.value = priority
    return {(table.sheet_name, profile_cell.coordinate), (table.sheet_name, priority_cell.coordinate)}


def build_review_output(source, output, owner=None, checkpoint=None, auditor=None, category="", session="top_merchants_review"):
    source_path = Path(source).resolve()
    output_path = Path(output).resolve()
    if source_path == output_path:
        raise ReviewWorkbookError("output must not overwrite source workbook")
    if not source_path.is_file():
        raise ReviewWorkbookError(f"source workbook does not exist: {source_path}")

    source_workbook = load_workbook(
        source_path,
        data_only=False,
        keep_vba=source_path.suffix.lower() == ".xlsm",
    )
    tables = discover_review_tables(source_workbook)
    selected_owner = resolve_owner(source_workbook, tables, owner)
    selection = select_review_rows(source_workbook, tables, selected_owner)
    existing_results = _completed_results(selection.completed)
    reusable_results = existing_results
    tasks = group_pending_rows(selection.pending)
    if category:
        tasks = [replace(task, category=task.category or category) for task in tasks]

    task_values = {}
    unresolved = []
    for task in tasks:
        key = normalize_shop_name(task.shop_name)
        workbook_key = (key, normalize_url(task.shop_url))
        if workbook_key in reusable_results:
            task_values[key] = reusable_results[workbook_key]
        else:
            unresolved.append(task)
    if unresolved and auditor is None:
        checkpoint_path = Path(checkpoint) if checkpoint else output_path.with_suffix(".review-checkpoint.json")
        browser = WebcliBrowserAdapter(session=session)
        audited_results = audit_queue(unresolved, browser, checkpoint_path)
        for task, result in zip(unresolved, audited_results, strict=True):
            task_values[normalize_shop_name(task.shop_name)] = _decision_values(result)
    elif unresolved:
        for task in unresolved:
            task_values[normalize_shop_name(task.shop_name)] = _decision_values(auditor.audit(task))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, output_path)
    workbook = load_workbook(
        output_path,
        data_only=False,
        keep_vba=output_path.suffix.lower() == ".xlsm",
    )
    output_tables = discover_review_tables(workbook)
    tables_by_sheet = _table_by_sheet(output_tables)
    allowed_cells = _phone_cells_to_text(workbook, output_tables)
    reused_rows = 0
    audited_shops = 0

    for task in tasks:
        key = normalize_shop_name(task.shop_name)
        values = task_values[key]
        if (key, normalize_url(task.shop_url)) in reusable_results:
            reused_rows += len(task.source_rows)
        else:
            audited_shops += 1
        for sheet_name, row_number in task.source_rows:
            allowed_cells.update(
                _write_result(workbook, tables_by_sheet[sheet_name], row_number, values[0], values[1])
            )

    workbook.save(output_path)
    if not zipfile.is_zipfile(output_path):
        raise ReviewWorkbookError("output is not a valid XLSX ZIP package")
    load_workbook(output_path, data_only=False, keep_vba=output_path.suffix.lower() == ".xlsm").close()
    differences = compare_workbooks(source_path, output_path, allowed_cells)
    if differences:
        raise ReviewWorkbookError("unexpected workbook changes: " + "; ".join(differences[:10]))
    return ReviewRunSummary(
        output_path=str(output_path),
        owner=selected_owner,
        pending_rows=len(selection.pending),
        unique_tasks=len(tasks),
        reused_rows=reused_rows,
        audited_shops=audited_shops,
        inconsistent_rows=tuple((row.sheet_name, row.row_number) for row in selection.inconsistent),
    )


def main(argv=None):
    parser = argparse.ArgumentParser(description="审核工作簿中的淘宝/天猫店铺并另存结果")
    parser.add_argument("source")
    parser.add_argument("--output", required=True)
    parser.add_argument("--owner")
    parser.add_argument("--checkpoint")
    parser.add_argument("--category", default="")
    parser.add_argument("--session", default="top_merchants_review")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args(argv)

    if args.dry_run:
        print(json.dumps(plan_review_workbook(args.source, args.owner, args.category), ensure_ascii=False, indent=2))
        return 0
    try:
        summary = build_review_output(
            args.source,
            args.output,
            owner=args.owner,
            checkpoint=args.checkpoint,
            category=args.category,
            session=args.session,
        )
    except AuditPaused as error:
        print(json.dumps({"status": "paused", "reason": error.reason, "checkpoint": str(error.checkpoint_path)}, ensure_ascii=False))
        return 2
    print(json.dumps({"status": "completed", **summary.__dict__}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
